from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import sqlite3, os, uuid, io, zipfile, json
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import fitz  # PyMuPDF for template overlay

app = Flask(__name__)
CORS(app)
DB = 'bupreme.db'

def init_db():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (id TEXT PRIMARY KEY, name TEXT, email TEXT UNIQUE, password TEXT, phone TEXT, plan TEXT, created TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS generations (id TEXT, user_id TEXT, players INTEGER, created TEXT, order_id TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS notifications (id TEXT, user_email TEXT, type TEXT, title TEXT, body TEXT, created TEXT, read INTEGER DEFAULT 0)''')
    try:
        c.execute("INSERT INTO users VALUES (?,?,?,?,?,?,?)", 
                 ('admin','Admin','dhub0005@gmail.com', generate_password_hash('dhub@0005'), '', 'admin', datetime.now().isoformat()))
    except: pass
    conn.commit(); conn.close()

init_db()

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/signup', methods=['POST'])
def signup():
    d = request.json
    conn = sqlite3.connect(DB)
    try:
        uid = str(uuid.uuid4())
        conn.execute("INSERT INTO users VALUES (?,?,?,?,?,?,?)",
                    (uid, d['name'], d['email'], generate_password_hash(d['password']), d.get('phone',''), 'trial', datetime.now().isoformat()))
        conn.commit()
        return jsonify({'ok':True, 'user':{'id':uid,'name':d['name'],'email':d['email'],'plan':'trial'}})
    except:
        return jsonify({'ok':False,'error':'Email exists'}), 400
    finally: conn.close()

@app.route('/api/login', methods=['POST'])
def login():
    d = request.json
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT id,name,email,password,plan FROM users WHERE email=?", (d['email'],))
    r = c.fetchone(); conn.close()
    if r and check_password_hash(r[3], d['password']):
        return jsonify({'ok':True,'user':{'id':r[0],'name':r[1],'email':r[2],'plan':r[4]}})
    return jsonify({'ok':False}), 401

@app.route('/api/parse_excel', methods=['POST'])
def parse_excel():
    f = request.files['file']
    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            vals = [str(c or '').lower() for c in row]
            if any('name' in v for v in vals) and any('no' in v or 'number' in v for v in vals):
                header_idx = i; break
        headers = [str(c or '').strip().lower() for c in rows[header_idx]]
        def find(names):
            for n in names:
                for idx,h in enumerate(headers):
                    if n in h.replace(' ',''):
                        return idx
            return None
        ncol = find(['name','player','fullname'])
        numcol = find(['number','no','jersey','num'])
        scol = find(['size','chest','sz'])
        players = []
        for row in rows[header_idx+1:]:
            if not row: continue
            name = str(row[ncol] or '').strip() if ncol is not None and ncol < len(row) else ''
            num = str(row[numcol] or '').strip() if numcol is not None and numcol < len(row) else ''
            size = str(row[scol] or '').strip() if scol is not None and scol < len(row) else ''
            if name and num and name.lower()!='none':
                players.append({'name':name.upper(),'number':num,'size':size})
        return jsonify({'ok':True,'players':players})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 400

@app.route('/api/generate', methods=['POST'])
def generate():
    data = request.form
    players = json.loads(data.get('players','[]'))
    user_id = data.get('user_id','anon')
    order_id = data.get('order_id','')
    
    templates = []
    for key in request.files:
        if key.startswith('template'):
            templates.append(request.files[key].read())
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in players:
            # If templates uploaded, overlay on each
            if templates:
                for idx, tpl_bytes in enumerate(templates):
                    try:
                        doc = fitz.open(stream=tpl_bytes, filetype="pdf")
                        page = doc[0]
                        # overlay name and number - center positions (adjust as needed)
                        rect_name = fitz.Rect(0, 100, page.rect.width, 200)
                        rect_num = fitz.Rect(0, page.rect.height/2 - 100, page.rect.width, page.rect.height/2 + 100)
                        page.insert_textbox(rect_name, p['name'], fontsize=48, align=1, fontname="helv")
                        page.insert_textbox(rect_num, p['number'], fontsize=180, align=1, fontname="helv")
                        pdf_bytes = doc.tobytes()
                        zf.writestr(f"{p['name']}_{p['number']}_template{idx+1}.pdf", pdf_bytes)
                        doc.close()
                    except Exception as e:
                        # fallback to simple
                        pass
            else:
                # simple fallback (original)
                img = Image.new('RGB', (2480,3508), 'white')
                d = ImageDraw.Draw(img)
                d.rectangle([0,0,2480,80], fill='#2563eb')
                d.rectangle([0,3428,2480,3508], fill='#2563eb')
                try:
                    f1 = ImageFont.truetype("DejaVuSans.ttf", 180)
                    f2 = ImageFont.truetype("DejaVuSans-Bold.ttf", 600)
                except:
                    f1 = ImageFont.load_default(); f2 = ImageFont.load_default()
                d.text((1240,600), p['name'], fill='black', font=f1, anchor='mm')
                d.text((1240,1750), p['number'], fill='black', font=f2, anchor='mm')
                d.text((2200,3300), f"Size: {p['size']}", fill='black', font=f1, anchor='rm')
                pdf_bytes = io.BytesIO()
                img.save(pdf_bytes, format='PDF', resolution=300)
                zf.writestr(f"{p['name']}_{p['number']}.pdf", pdf_bytes.getvalue())
    
    conn = sqlite3.connect(DB)
    conn.execute("INSERT INTO generations VALUES (?,?,?,?,?)", (str(uuid.uuid4()), user_id, len(players), datetime.now().isoformat(), order_id))
    conn.commit(); conn.close()
    
    zip_buffer.seek(0)
    return send_file(zip_buffer, mimetype='application/zip', as_attachment=True, download_name=f'{order_id or "jerseys"}.zip')

@app.route('/api/send_message', methods=['POST'])
def send_message():
    d = request.json
    conn = sqlite3.connect(DB)
    conn.execute("INSERT INTO notifications VALUES (?,?,?,?,?,?,?)",
                (str(uuid.uuid4()), d['email'], d['type'], d['title'], d['body'], datetime.now().isoformat(), 0))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/notifications')
def get_notifications():
    email = request.args.get('email','')
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    if email == 'dhub0005@gmail.com':  # admin sees all
        c.execute("SELECT type,title,body,created FROM notifications ORDER BY created DESC LIMIT 50")
    else:
        c.execute("SELECT type,title,body,created FROM notifications WHERE user_email=? ORDER BY created DESC LIMIT 50", (email,))
    rows = [{'type':r[0],'title':r[1],'body':r[2],'created':r[3]} for r in c.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/stats')
def stats():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT COALESCE(SUM(players),0) FROM generations"); pdfs = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*) FROM users WHERE id!='admin'"); clients = c.fetchone()[0] or 0
    conn.close()
    return jsonify({'pdfs':pdfs,'clients':clients,'revenue':0})

@app.route('/download-source')
def download_source():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for path in ['app.py', 'requirements.txt', 'README.md']:
            if os.path.exists(path): zf.write(path)
        tpl = os.path.join('templates', 'index.html')
        if os.path.exists(tpl): zf.write(tpl, 'templates/index.html')
    buf.seek(0)
    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name='jersey-v12-source.zip')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
